import pandas as pd
import numpy as np
import ast
import os
import re
import json
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from sklearn.linear_model import SGDClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import LinearSVC
from sklearn.multiclass import OneVsRestClassifier
from PyPDF2 import PdfReader

import pickle
from datetime import datetime


multilabel = MultiLabelBinarizer()
allCV={}
listOfTitles=[]

def openLists():
  textList=[]
  kategori1 = open("kategori1.txt","r",encoding='utf-8')
  for line in kategori1:
      textList.append(line.strip())
  kategori1.close()
  return textList

def stopList():
  myfile=open("stoplista.txt","r",encoding='utf-8')
  stoplistVectorizer=TfidfVectorizer(lowercase=True)
  stopList=[]
  for line in myfile:
      stopList.append(line)
  myfile.close() 

  stoplistVectorizer.fit(stopList)
  tokenizedStopWords=stoplistVectorizer.get_feature_names_out()
  tokenizedStopWords=tokenizedStopWords.tolist()
  return tokenizedStopWords

# Rensa och konvertera data till listor av strängar
def clean_and_convert_to_list(text):
    # Kontrollerar först om 'text' är en sträng
    if isinstance(text, str):
        if text.startswith('[') and text.endswith(']'):
            text = text[1:-1]  # Tar bort hakparenteserna
        return [item.strip() for item in text.split(',')]
    else:
        # Returnerar en tom lista om text inte är en sträng
        return []
# Applicera funktionen på Attribut-kolumnen

def readPDFCV(files, pdfFilePath):
  print("------------------------------Läsa CV--------------------")
  if files.endswith('.pdf'):
      pdfPath =pdfFilePath+files  
      print(f"Namn på fil {files}")
      #pdf_path = pdf_path+'.pdf'
      wholeDocument=""
      print(pdfPath)
      reader = PdfReader(pdfPath)
      #page = reader.pages[1]
      for sida in reader.pages:
        wholeDocument+=sida.extract_text()+'\n'
  return wholeDocument


###ÖPPNA EXCELDOKUMENT
dataframe=pd.read_excel('testfil2.xlsx')
#ast.literal_eval(dataframe['Attribut'].iloc[0])
for i in dataframe['Yrkestitel']:
   listOfTitles.append(dataframe['Yrkestitel'])
##print(f"LISTOFTITLES :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::{listOfTitles}")
  
dataframe['Attribut'] = dataframe['Attribut'].apply(clean_and_convert_to_list)
excelTraningDataProperties= load_workbook('testfil2.xlsx')

###VECTORIZE/TOKENIZ
def train_model():
  print("----------------Bygga model------------------")
  y = multilabel.fit_transform(dataframe['Attribut'])
  dataframe['Attribut'] = dataframe['Attribut'].apply(lambda x: ' '.join(x)) ##Konverterar till en sträng
  tfidf = TfidfVectorizer(analyzer='word', ngram_range=(1,2), stop_words=stopList(), lowercase=True, )

  X = tfidf.fit_transform(dataframe['Yrkestitel'])
  X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.2, random_state = 0) 

  ###TRÄNA MODELL
  
  
  linres=LinearSVC(C=2,penalty='l1', dual=False, class_weight="balanced",max_iter=50000)     ##BARA 3 FEL!!!!!!!!
  clf = OneVsRestClassifier(linres)
  fileName="model.sav"
  pickle.dump(clf,open(fileName,'wb'))
  clf.fit(X_train, y_train)
  
  clf=pickle.load(open(fileName, 'rb'))
  clf.fit(X_train, y_train)
  print(clf.score(X_train, y_train))
  return tfidf, clf


def cheackForChanges():
    traningData="testfil2.xlsx"
    filePath="./"+traningData
    metaData="./metaData.json"

    excelTraningDataProperties=os.stat(filePath)
    jsonExcelProperties=json.dumps(excelTraningDataProperties.st_mtime)
    with open(metaData,'r') as data:
        oldProperties=data.read()
    if oldProperties!=str(jsonExcelProperties):
       with open(metaData, 'w') as outData:
          outData.write(jsonExcelProperties)
       return True
    return False
    #excelTraningDataProperties=json.dumps(load_workbook(traningData+'.xlsx'))



def j_score(y_true, y_pred):
  jaccard = np.minimum(y_true, y_pred).sum(axis = 1)/np.maximum(y_true, y_pred).sum(axis = 1)
  return jaccard.mean()*100


def run_model(tfidf, clf):
  if excelTraningDataProperties!=load_workbook("testfil2.xlsx"):
    print("Modellen är inte uppdaterad, vänligen träna om modellen")

  pdfFilePath="./uploads/"
  for files in os.listdir(pdfFilePath):
      numberOfExclusiveHitsForProcent=0
      CV=readPDFCV(files,pdfFilePath)
      x=CV.split()
      xt=tfidf.transform(x)
      attributesFromCV=multilabel.inverse_transform(clf.predict(xt))
      wantedAttributes=['affärsmässig','numerisk analytisk förmåga','kvalitetsmedveten','språklig analytisk förmåga']
      #listOfAttributeCleaned= [str(t) for t in attributesFromCV if t]
      realCleanList=[]
      score=0
      setOfAttributes=set()
      listOfAttributeCleaned=attributesFromCV
      for cleanAttributes in listOfAttributeCleaned:
        for tuples in cleanAttributes:
              realCleanList.append(tuples)
              setOfAttributes.add(tuples)
              
      print("---------------------------------------------------------------------------------------------")
      print("---------------------------------------------------------------------------------------------")

      for i in setOfAttributes:
        if i in wantedAttributes:
          numberOfExclusiveHitsForProcent=numberOfExclusiveHitsForProcent+1

      wantedAttributeScoringDict = {}
      scoringDict = {}

      for attribut in realCleanList:
        scoringDict[attribut] = scoringDict.get(attribut, 0) + 1
        if attribut in wantedAttributes:
            wantedAttributeScoringDict[attribut] = wantedAttributeScoringDict.get(attribut, 0) + 1
            score=score+1

      print("\n---------------------------------------------------------------------------------------------")
      print("Score per attribute: ", scoringDict)
      print("Score per wanted attribute: ", wantedAttributeScoringDict)
      print("---------------------------------------------------------------------------------------------\n")

      procentOfAttributes=(numberOfExclusiveHitsForProcent/len(wantedAttributes))*100
      print(len(wantedAttributes))
      print(numberOfExclusiveHitsForProcent)

      ## len(realCleanList) ger väl hur många attribut som uppfylls överhuvudtaget, inte nödvändigtvis att de är önskade?
      print(f"Antalet gånger ett attributet uppfylls {wantedAttributes} är {len(realCleanList)}")
      print(f"Andel av attributen som uppfylls {procentOfAttributes} %")
      thisCV={}
      thisCV={"Score": len(realCleanList),"PercentScore":procentOfAttributes}
      allCV.update({files:thisCV})
#print(allCV)
jsonDict=json.dumps(allCV)
print(jsonDict)

if __name__ == '__main__':
  tfidf, clf = train_model()
  while(input("Avsluta? (j) ")!='j'):
    run_model(tfidf, clf)
    metadata='metaData'
    
    print(cheackForChanges())